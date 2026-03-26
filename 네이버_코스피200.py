import re
import sys

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QApplication,
    QFileDialog,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
)

ENTRY_URL = "https://finance.naver.com/sise/entryJongmok.naver"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}
COL_HEADERS = ["No", "페이지", "종목별", "현재가", "전일비", "등락률", "거래량", "거래대금(백만)", "시가총액(억)"]


# ──────────────────────────────────────────────
#  크롤링 로직 (순수 함수)
# ──────────────────────────────────────────────

def get_total_pages():
    res = requests.get(ENTRY_URL, params={"type": "KPI200", "page": 1},
                       headers=HEADERS, timeout=10)
    res.raise_for_status()
    res.encoding = res.apparent_encoding
    soup = BeautifulSoup(res.text, "html.parser")

    pager = soup.select("td.pgRR a, .pgRR a")
    if pager:
        m = re.search(r"page=(\d+)", pager[-1].get("href", ""))
        if m:
            return int(m.group(1))

    numbers = [
        int(m.group(1))
        for a in soup.select(".pgNav a, table.pgNav a")
        if (m := re.search(r"page=(\d+)", a.get("href", "")))
    ]
    return max(numbers) if numbers else 20


def crawl_page(page: int):
    res = requests.get(ENTRY_URL, params={"type": "KPI200", "page": page},
                       headers=HEADERS, timeout=10)
    res.raise_for_status()
    res.encoding = res.apparent_encoding
    soup = BeautifulSoup(res.text, "html.parser")

    table = soup.select_one("table.type_1")
    if table is None:
        raise ValueError("편입종목상위 테이블을 찾지 못했습니다.")

    result = []
    for row in table.select("tr"):
        tds = row.find_all("td")
        if len(tds) != 7:
            continue
        cols = [td.get_text(" ", strip=True) for td in tds]
        if not cols[0]:
            continue
        result.append({
            "종목별": cols[0],
            "현재가": cols[1],
            "전일비": cols[2],
            "등락률": cols[3],
            "거래량": cols[4],
            "거래대금(백만)": cols[5],
            "시가총액(억)": cols[6],
        })
    return result


# ──────────────────────────────────────────────
#  백그라운드 크롤링 스레드
# ──────────────────────────────────────────────

class CrawlWorker(QThread):
    # 시그널 정의
    progress = pyqtSignal(int, int)          # (현재 페이지, 전체 페이지)
    item_ready = pyqtSignal(dict)            # 수집된 한 건
    status_msg = pyqtSignal(str)             # 상태 메시지
    finished_ok = pyqtSignal(int)            # 완료 (총 수집 건수)
    error = pyqtSignal(str)                  # 오류 메시지

    MAX_ITEMS = 200

    def run(self):
        try:
            self.status_msg.emit("전체 페이지 수 확인 중...")
            total_pages = get_total_pages()
            self.status_msg.emit(f"전체 페이지 수: {total_pages}")

            count = 0
            for page in range(1, total_pages + 1):
                self.progress.emit(page, total_pages)
                self.status_msg.emit(f"{page}/{total_pages} 페이지 수집 중...")

                items = crawl_page(page)
                if not items:
                    self.status_msg.emit(f"{page}페이지 데이터 없음 — 종료")
                    break

                for item in items:
                    item["페이지"] = page
                    self.item_ready.emit(item)
                    count += 1
                    if count >= self.MAX_ITEMS:
                        break

                if count >= self.MAX_ITEMS:
                    break

            self.finished_ok.emit(count)

        except Exception as exc:
            self.error.emit(str(exc))


# ──────────────────────────────────────────────
#  메인 윈도우
# ──────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("네이버 코스피200 편입종목 크롤러")
        self.resize(1000, 640)
        self._data: list[dict] = []
        self._worker: CrawlWorker | None = None
        self._setup_ui()

    # ── UI 구성 ──────────────────────────────

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(8)
        root.setContentsMargins(10, 10, 10, 10)

        # 버튼 영역
        btn_row = QHBoxLayout()
        self.btn_crawl = QPushButton("크롤링 시작")
        self.btn_crawl.setFixedHeight(36)
        self.btn_crawl.clicked.connect(self._on_crawl)

        self.btn_save = QPushButton("엑셀 저장")
        self.btn_save.setFixedHeight(36)
        self.btn_save.setEnabled(False)
        self.btn_save.clicked.connect(self._on_save)

        btn_row.addWidget(self.btn_crawl)
        btn_row.addWidget(self.btn_save)
        btn_row.addStretch()
        root.addLayout(btn_row)

        # 진행 바
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        root.addWidget(self.progress_bar)

        # 상태 레이블
        self.lbl_status = QLabel("대기 중")
        root.addWidget(self.lbl_status)

        # 테이블
        self.table = QTableWidget(0, len(COL_HEADERS))
        self.table.setHorizontalHeaderLabels(COL_HEADERS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        root.addWidget(self.table)

    # ── 슬롯 ─────────────────────────────────

    def _on_crawl(self):
        # 초기화
        self._data.clear()
        self.table.setRowCount(0)
        self.btn_save.setEnabled(False)
        self.btn_crawl.setEnabled(False)
        self.progress_bar.setValue(0)
        self.lbl_status.setText("크롤링 시작...")

        self._worker = CrawlWorker()
        self._worker.progress.connect(self._on_progress)
        self._worker.item_ready.connect(self._on_item_ready)
        self._worker.status_msg.connect(self.lbl_status.setText)
        self._worker.finished_ok.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_progress(self, current: int, total: int):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.progress_bar.setFormat(f"{current} / {total} 페이지")

    def _on_item_ready(self, item: dict):
        self._data.append(item)
        row_idx = self.table.rowCount()
        self.table.insertRow(row_idx)

        values = [
            str(row_idx + 1),
            str(item.get("페이지", "")),
            item.get("종목별", ""),
            item.get("현재가", ""),
            item.get("전일비", ""),
            item.get("등락률", ""),
            item.get("거래량", ""),
            item.get("거래대금(백만)", ""),
            item.get("시가총액(억)", ""),
        ]

        for col, val in enumerate(values):
            cell = QTableWidgetItem(val)
            cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            # 등락률 색상 (▲ 빨강 / ▼ 파랑)
            if col == 5:
                if "+" in val or "▲" in val:
                    cell.setForeground(QColor("#cc0000"))
                elif "-" in val or "▼" in val:
                    cell.setForeground(QColor("#0055cc"))

            self.table.setItem(row_idx, col, cell)

        # 새 행이 보이도록 스크롤
        self.table.scrollToBottom()

    def _on_finished(self, count: int):
        self.lbl_status.setText(f"크롤링 완료 — 총 {count}건 수집")
        self.btn_crawl.setEnabled(True)
        self.btn_save.setEnabled(count > 0)
        self.progress_bar.setFormat(f"완료 ({count}건)")

    def _on_error(self, msg: str):
        self.lbl_status.setText(f"오류: {msg}")
        self.btn_crawl.setEnabled(True)
        QMessageBox.critical(self, "크롤링 오류", msg)

    def _on_save(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "엑셀 파일 저장", "kospi200.xlsx",
            "Excel 파일 (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "KOSPI200"

            col_names = ["No", "페이지", "종목별", "현재가", "전일비", "등락률",
                         "거래량", "거래대금(백만)", "시가총액(억)"]

            # ── 헤더 스타일 정의 ──────────────────
            header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            header_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="AAAAAA")
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # ── 헤더 행 작성 ─────────────────────
            ws.append(col_names)
            for col_idx, _ in enumerate(col_names, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = cell_border

            # ── 데이터 행 작성 ───────────────────
            # 등락률 색상 구분용
            red_font = Font(name="맑은 고딕", color="CC0000", size=10)
            blue_font = Font(name="맑은 고딕", color="0055CC", size=10)
            normal_font = Font(name="맑은 고딕", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")

            for i, row in enumerate(self._data, 2):  # 2행부터 (1행=헤더)
                rate = row.get("등락률", "")
                values = [
                    i - 1,
                    row.get("페이지", ""),
                    row.get("종목별", ""),
                    row.get("현재가", ""),
                    row.get("전일비", ""),
                    rate,
                    row.get("거래량", ""),
                    row.get("거래대금(백만)", ""),
                    row.get("시가총액(억)", ""),
                ]
                # 홀짝 행 배경색
                row_fill = PatternFill(fill_type="solid",
                                       fgColor="EBF3FF" if i % 2 == 0 else "FFFFFF")

                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=i, column=col_idx, value=val)
                    cell.border = cell_border
                    cell.fill = row_fill

                    # 등락률 색상
                    if col_idx == 6:
                        if "+" in str(val) or "▲" in str(val):
                            cell.font = red_font
                        elif "-" in str(val) or "▼" in str(val):
                            cell.font = blue_font
                        else:
                            cell.font = normal_font
                        cell.alignment = right_align
                    elif col_idx in (1, 2):  # No, 페이지
                        cell.font = normal_font
                        cell.alignment = center_align
                    elif col_idx == 3:       # 종목별
                        cell.font = normal_font
                        cell.alignment = center_align
                    else:                    # 숫자 컬럼
                        cell.font = normal_font
                        cell.alignment = right_align

            # ── 열 너비 자동 조정 ────────────────
            col_widths = [6, 8, 16, 12, 12, 10, 14, 16, 14]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # ── 헤더 행 높이 & 틀 고정 ──────────
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"  # 헤더 행 고정

            wb.save(file_path)
            QMessageBox.information(
                self, "저장 완료",
                f"총 {len(self._data)}건이 저장되었습니다.\n{file_path}"
            )
        except Exception as exc:
            QMessageBox.critical(self, "저장 오류", str(exc))


# ──────────────────────────────────────────────
#  진입점
# ──────────────────────────────────────────────

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())